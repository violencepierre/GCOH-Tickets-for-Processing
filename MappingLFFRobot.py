###PCM window should be in the positive coords####


import os
import pyautogui as pya
import openpyxl as xl
import time
import threading
import pyperclip
import keyboard


global child
global parent
global region

def init(child,parent):
    global findall
    global searchfield
    global searchterm
    global findbutton
    global edit
    global cut
    global paste

    pyperclip.copy(child)
    pyperclip.waitForPaste()
    
    #look for the findall, edit, searchfield, and the find button
    print('Analyzing screen...')
    region=(0, 0, 1920, 1080)

    findall=None
    while findall is None:
        try:
            findall=pya.locateOnScreen(r'GCOH Resources\FindAllButton.png',grayscale=True,region=region,confidence=.70)
            ##print('find is found in ' + str(pya.center(findall)))
        except Exception as e:
            print('waiting waiting waiting.....')
    searchterm=None
    while searchterm is None:
        try:
            searchterm=pya.locateOnScreen(r'GCOH Resources\SearchTerm.png',grayscale=True,region=region,confidence=.70)

            x, y = pya.center(searchterm)
            x,y=x+100,y
            search=(x,y) ###searchfield
        except Exception as e:
            print('waiting waiting waiting.....')

    pya.moveTo(search)
    time.sleep(.7)
    pya.doubleClick(search)
    pya.doubleClick(search)
    pya.doubleClick(search)
    pya.hotkey("ctrl","a")
    pya.hotkey("ctrl","a")

    time.sleep(0.7)
###pya.doubleClick(search)
    ##pya.typewrite(child,interval=0.20)
    keyboard.send('ctrl+v')
    


    time.sleep(0.5)

    pya.click(findall)

    hitlist = None

    while hitlist is None:
        try:
            hitlist = pya.locateOnScreen(r'GCOH Resources\HitList Message.png', grayscale=True, region=region,confidence=.70)
        except Exception as e:
            print('waiting waiting waiting.....')


    hitlist=pya.center(hitlist)
    x,y=hitlist
    hitlist=x+10,y+80

    pya.click(hitlist)

    time.sleep(2)

    editbutton = None
    while editbutton is None:
        try:
            editbutton = pya.locateOnScreen(r'GCOH Resources\editbutton.png', grayscale=True, region=region,
                                            confidence=.70)
        except Exception as e:
            print('waiting waiting waiting.....')

    editbutton = pya.center(editbutton)
    x, y = editbutton

    pya.click(editbutton)

    cutcommand = None
    while cutcommand is None:
        try:
            cutcommand = pya.locateOnScreen(r'GCOH Resources\cutcommand.png', grayscale=True, region=region,
                                            confidence=.70)
        except Exception as e:
            print('waiting waiting waiting.....')

    pya.click(cutcommand)

    findall=None
    while findall is None:
        try:
            findall=pya.locateOnScreen(r'GCOH Resources\FindAllButton.png',grayscale=True,region=region,confidence=.70)
            ##print('find is found in ' + str(pya.center(findall)))
        except Exception as e:
            print('waiting waiting waiting.....')

    searchterm=None
    while searchterm is None:
        try:
            searchterm=pya.locateOnScreen(r'GCOH Resources\SearchTerm.png',grayscale=True,region=region,confidence=.70)

            x, y = pya.center(searchterm)
            x,y=x+100,y
            search=(x,y) ###searchfield
        except Exception as e:
            print('waiting waiting waiting.....')

    pyperclip.copy(parent)
    pyperclip.waitForPaste()

    pya.moveTo(search)
    time.sleep(0.7)
    pya.doubleClick(search)
    pya.doubleClick(search)
    pya.doubleClick(search)
    time.sleep(0.7)
    pya.press('delete')
    pya.press('delete')
    pya.hotkey("ctrl","a")
    pya.hotkey("ctrl","a")
    time.sleep(0.7)
    ###pya.typewrite(parent,interval=0.20)
    keyboard.send('ctrl+v')
    time.sleep(0.7)
    pya.click(findall)






    hitlist = None

    while hitlist is None:
        try:
            hitlist = pya.locateOnScreen(r'GCOH Resources\HitList Message.png', grayscale=True, region=region,confidence=.70)
        except Exception as e:
            print('waiting waiting waiting.....')


    hitlist=pya.center(hitlist)
    x,y=hitlist
    hitlist=x+10,y+80

    pya.click(hitlist)

    time.sleep(2)

    editbutton = None
    while editbutton is None:
        try:
            editbutton = pya.locateOnScreen(r'GCOH Resources\editbutton.png', grayscale=True, region=region,
                                            confidence=.70)
        except Exception as e:
            print('waiting waiting waiting.....')

    editbutton = pya.center(editbutton)
    x, y = editbutton

    pya.click(editbutton)

    pasteatlowerlevel = None
    while pasteatlowerlevel is None:
        try:
            pasteatlowerlevel = pya.locateOnScreen(r'GCOH Resources\pasteatlowerlevel.png', grayscale=True, region=region,confidence=.70)
        except Exception as e:
            print('waiting waiting waiting.....')

    pya.click(pasteatlowerlevel)
    time.sleep(2)

def mainloop():

    wb=xl.load_workbook(filename=r'GCOH Resources/GCOHPost.xlsx')
    ws=wb["MDGF"]
    lrow=len(ws['A'])
    i=2

    while i < lrow+1:


        child=ws['A' + str(i)].value
        parent=ws['B' + str(i)].value
        os.system("cls")
        print('Working on row:' + str(i))
        print(child + " goes to " + parent)
        i=i+1
        print('Currently remaining objects: ' + str(lrow - 1 - (i-2)) + ' (aside from currently being processed)') 
        

        init(child,parent)

    print("Task completed successfully")



##dream()
